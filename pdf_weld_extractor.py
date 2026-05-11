"""
技术路线二：使用多模态大模型从 PDF 加工图中提取焊缝数据
策略：自动识别截面视图(A-A/B-B/...) → 裁剪 → 高分辨率PNG → 逐视图调用LLM
依赖：pip install pymupdf openai openpyxl
"""

import os
import re
import base64
import json
import glob
import time
import traceback

import fitz                     # PyMuPDF
from openai import OpenAI
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import openpyxl.utils

# ── API 配置 ──────────────────────────────────────────────────────────────────
API_KEY         = "sk-URBynpfNYkkj3kewnxeA"
CUSTOM_BASE_URL = "http://172.18.162.10:8002/v1"
MODEL_NAME      = "Qwen/Qwen3.5-27B"

# ── 路径 ──────────────────────────────────────────────────────────────────────
FOLDER = r"c:\Users\hp\OneDrive\Desktop\dxf\hanf"
OUTPUT = os.path.join(FOLDER, "焊缝统计_llm.xlsx")

# ── 渲染参数 ──────────────────────────────────────────────────────────────────
CROP_DPI    = 400   # 裁剪视图的渲染DPI，高DPI确保小字/符号清晰
VIEW_MARGIN = 45    # 视图裁剪外扩边距（PDF点，约1.5cm）
RETRY_TIMES = 3
RETRY_DELAY = 3

# ─────────────────────────────────────────────────────────────────────────────
# Prompt
# ─────────────────────────────────────────────────────────────────────────────
SYSTEM_PROMPT = (
    "你是专业的钢结构工程识图助手，擅长从加工图中读取焊缝符号和标注。"
    "请严格按照用户要求的JSON格式输出，不要包含任何解释文字。"
)

# 主视图（立面图）专用提示词
MAIN_PROMPT_TMPL = """\
这是钢结构构件 {comp} 加工图的主立面视图局部（比例1:10）。
请找出图中**所有焊缝标注符号**，提取每条焊缝信息。

【焊缝符号识别规则】
- 焊缝符号：带箭头的引线 + 参考线，参考线上/下有实心三角（角焊缝），三角旁数字=hf（mm）
- 参考线上下均有三角=双面焊，Above和Below分别记一行
- 引线箭头所指位置附近的零件编号（pXXX或{comp}）即为连接零件

【重要提示】
- 图中若存在"pXXX (N/S)"标注（如p52 N/S），表示两侧各有该零件，不是焊缝的直接连接对象
- 梁端部（图左侧）的焊缝通常连接端板（end plate），若标注"p118"则part2=p118
- 腹板/翼缘与不同零件的焊缝要分别记录

【输出JSON数组】
字段：component("{comp}"), position("Above"/"Below"), hf(整数),
      length_mm(0，除非图中有明确长度标注), annotation(""), part1("{comp}"), part2(零件编号)

只输出JSON，不要任何解释。
"""

VIEW_PROMPT_TMPL = """\
这是钢结构构件 {comp} 加工图中的截面视图 "{section}"（比例1:10）的高分辨率PNG图像。
请仔细找出图中所有焊缝标注，提取每条焊缝信息。

【焊缝符号识别】
- 焊缝符号：箭头+参考线，参考线上方或下方有实心三角形（角焊缝符号）
- 三角形旁的数字 = 焊脚尺寸 hf（单位mm，直接读图中数字）
- 参考线尾端标注的数字 = 焊缝长度（单位mm，直接读数字）
- 参考线上下各有三角 = 双面焊，Above和Below各记一行
- 标注含"3 SIDES"或"三面" = 三面围焊，每条边单独一行
- 引线箭头指向处附近的字符 = 零件编号（如p52、p118、{comp}）

【输出JSON数组，每条焊缝一个对象】
字段：component, position("Above"/"Below"), hf(整数), length_mm(数字),
      annotation("3 SIDES"或""), part1(零件1编号), part2(零件2编号)

输出示例：
[
  {{"component":"{comp}","position":"Above","hf":12,"length_mm":630,"annotation":"","part1":"{comp}","part2":"p118"}},
  {{"component":"{comp}","position":"Below","hf":12,"length_mm":630,"annotation":"","part1":"{comp}","part2":"p118"}}
]
只输出JSON，不要任何解释。
"""

# ─────────────────────────────────────────────────────────────────────────────
# 图像处理
# ─────────────────────────────────────────────────────────────────────────────
client = OpenAI(api_key=API_KEY, base_url=CUSTOM_BASE_URL)


def crop_view_to_png_b64(page: fitz.Page, rect: fitz.Rect,
                          dpi: int = CROP_DPI) -> str:
    """
    裁剪页面指定矩形区域，渲染为高分辨率PNG并返回base64字符串。
    PNG（无损）比JPEG更适合工程线条图，无压缩模糊。
    """
    mat = fitz.Matrix(dpi / 72.0, dpi / 72.0)
    pix = page.get_pixmap(matrix=mat, clip=rect, colorspace=fitz.csRGB)
    return base64.b64encode(pix.tobytes("png")).decode("utf-8")


def detect_section_views(page: fitz.Page, margin: float = VIEW_MARGIN):
    """
    扫描页面文字，自动识别截面视图标题（如"A - A"、"B - B"）。
    对每个标题，收集其下方附近的文字范围，确定视图裁剪矩形。
    返回：[(视图标签, fitz.Rect), ...]
    若未找到截面标题，返回整页作为单一视图。
    """
    spans = []
    for block in page.get_text("dict")["blocks"]:
        for line in block.get("lines", []):
            for sp in line.get("spans", []):
                t = sp["text"].strip()
                if t:
                    spans.append({"text": t, "bbox": sp["bbox"]})

    # 截面标题格式: "A - A", "B-B", "A–A" 等
    section_spans = [
        sp for sp in spans
        if re.match(r"^[A-Z]\s*[-\u2013\u2014]\s*[A-Z]$", sp["text"])
    ]

    if not section_spans:
        print("    未找到截面标题，使用整页")
        return [("全图", page.rect)]

    pw_val, ph_val = page.rect.width, page.rect.height
    # 估计标题栏起始 y（避免将图纸注释纳入裁剪）
    TITLE_BLOCK_Y = ph_val * 0.85   # 页面高度85%以下认为是标题栏/注释区

    SAME_COL_X = 300   # 两个标题 x 距离小于此值才视为同一列

    views = []
    for lbl_sp in section_spans:
        lx0, ly0, lx1, ly1 = lbl_sp["bbox"]
        lbl_cx = (lx0 + lx1) / 2

        # 找同列中位于此标题正下方的下一个标题，限制搜索高度
        col_below = [
            s for s in section_spans
            if s is not lbl_sp
            and abs((s["bbox"][0] + s["bbox"][2]) / 2 - lbl_cx) < SAME_COL_X
            and s["bbox"][1] > ly1
        ]
        if col_below:
            col_below.sort(key=lambda s: s["bbox"][1])
            next_y = col_below[0]["bbox"][1] - 10
        else:
            next_y = TITLE_BLOCK_Y

        max_search_y = min(next_y, ly0 + 320)   # 最多向下320pt

        # 左侧100pt、右侧250pt 搜索附近文字（左侧窄以避免主视图污染）
        nearby = [
            sp for sp in spans
            if (lbl_cx - 100) <= (sp["bbox"][0] + sp["bbox"][2]) / 2 <= (lbl_cx + 250)
            and (ly0 - 20) <= (sp["bbox"][1] + sp["bbox"][3]) / 2 <= max_search_y
        ]
        if not nearby:
            nearby = [lbl_sp]

        all_x = [sp["bbox"][0] for sp in nearby] + [sp["bbox"][2] for sp in nearby]
        all_y = [sp["bbox"][1] for sp in nearby] + [sp["bbox"][3] for sp in nearby]
        rx0 = max(0,             min(all_x) - margin)
        ry0 = max(0,             min(all_y) - margin)
        rx1 = min(pw_val,        max(all_x) + margin)
        ry1 = min(TITLE_BLOCK_Y, max(all_y) + margin + 20)
        # 强制最小尺寸：对称扩展，确保视图包含图形内容而不仅是标题文字
        MIN_W, MIN_H = 220, 240
        if rx1 - rx0 < MIN_W:
            extra = (MIN_W - (rx1 - rx0)) / 2
            rx0 = max(0,      rx0 - extra)
            rx1 = min(pw_val, rx0 + MIN_W)
        if ry1 - ry0 < MIN_H:
            ry1 = min(TITLE_BLOCK_Y, ry0 + MIN_H)
        rect = fitz.Rect(rx0, ry0, rx1, ry1)
        label = lbl_sp["text"].replace(" ", "")
        views.append((label, rect))
        print(f"    视图 {label}: x={rect.x0:.0f}-{rect.x1:.0f}, "
              f"y={rect.y0:.0f}-{rect.y1:.0f} "
              f"({rect.width:.0f}×{rect.height:.0f}pt)")

    return views


# ─────────────────────────────────────────────────────────────────────────────
# LLM 调用
# ─────────────────────────────────────────────────────────────────────────────
def extract_text_in_rect(page: fitz.Page, rect: fitz.Rect) -> str:
    """提取指定矩形区域内的文字及其坐标，供LLM参考（格式：(x,y) 文字）。"""
    items = []
    for block in page.get_text("dict")["blocks"]:
        for line in block.get("lines", []):
            for sp in line.get("spans", []):
                t = sp["text"].strip()
                if not t:
                    continue
                bx0, by0, bx1, by1 = sp["bbox"]
                cx, cy = (bx0 + bx1) / 2, (by0 + by1) / 2
                if rect.x0 <= cx <= rect.x1 and rect.y0 <= cy <= rect.y1:
                    items.append((round(cy, 1), round(cx, 1), t))
    items.sort()
    return "\n".join(f"({x:.0f},{y:.0f}) {txt}" for y, x, txt in items)


def call_llm(b64_png: str, comp: str, section: str = "detail",
             text_ctx: str = "") -> str:
    """调用多模态LLM（Qwen3，thinking模式已关闭），返回原始响应文本。"""
    if section == "主视图":
        prompt = MAIN_PROMPT_TMPL.format(comp=comp)
    else:
        prompt = VIEW_PROMPT_TMPL.format(comp=comp, section=section)
    if text_ctx:
        prompt += (
            "\n\n【视图区域内提取到的所有文字（坐标,内容）供参考】\n"
            + text_ctx
            + "\n（以上坐标与图像像素坐标对应，可辅助定位焊缝标注数值）"
        )
    messages = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {
            "role": "user",
            "content": [
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/png;base64,{b64_png}",
                        "detail": "high",
                    },
                },
                {"type": "text", "text": prompt},
            ],
        },
    ]
    for attempt in range(1, RETRY_TIMES + 1):
        try:
            resp = client.chat.completions.create(
                model=MODEL_NAME,
                messages=messages,
                temperature=0.05,
                max_tokens=8192,
                extra_body={"chat_template_kwargs": {"enable_thinking": False}},
            )
            content = resp.choices[0].message.content
            if content is None:
                choice = resp.choices[0]
                print(f"    警告: content=None, finish_reason={choice.finish_reason}")
                print(f"    完整choice: {choice}")
                return "[]"
            return content
        except Exception as e:
            print(f"    LLM调用失败（第{attempt}次）: {e}")
            traceback.print_exc()
            if attempt < RETRY_TIMES:
                time.sleep(RETRY_DELAY)
    return "[]"


# ─────────────────────────────────────────────────────────────────────────────
# 响应解析
# ─────────────────────────────────────────────────────────────────────────────
def parse_llm_response(text: str) -> list:
    """从LLM响应中健壮地提取JSON数组。"""
    # 优先提取 ```json ... ``` 块
    m = re.search(r"```(?:json)?\s*([\s\S]*?)```", text)
    candidate = m.group(1) if m else text

    if not m:
        m2 = re.search(r"(\[[\s\S]*\])", candidate)
        if m2:
            candidate = m2.group(1)

    try:
        data = json.loads(candidate.strip())
        if isinstance(data, list):
            return data
    except json.JSONDecodeError:
        pass

    # 逐行回退
    rows = []
    for line in text.splitlines():
        line = line.strip().rstrip(",")
        if line.startswith("{") and line.endswith("}"):
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError:
                pass
    return rows


def normalize_row(row: dict, comp: str) -> dict:
    """统一字段名、类型，兼容LLM的各种命名变体。"""
    aliases = {
        "weld_size": "hf", "size": "hf",
        "length": "length_mm", "weld_length": "length_mm",
        "remark": "annotation", "note": "annotation",
        "part_1": "part1", "part_2": "part2",
    }
    for alias, canon in aliases.items():
        if alias in row and canon not in row:
            row[canon] = row.pop(alias)

    row.setdefault("component", comp)
    row.setdefault("position", "Above")
    row.setdefault("annotation", "")
    row.setdefault("part1", comp)
    row.setdefault("part2", "")
    if row.get("annotation") is None:
        row["annotation"] = ""

    # 确保构件号在 part1
    if row.get("part2") == comp and row.get("part1") != comp:
        row["part1"], row["part2"] = row["part2"], row["part1"]

    for key in ("hf", "length_mm"):
        v = row.get(key)
        if v is None or str(v).lower() in ("null", "none", ""):
            row[key] = 0
        else:
            try:
                row[key] = float(v) if "." in str(v) else int(v)
            except (ValueError, TypeError):
                row[key] = 0
    return row


# ─────────────────────────────────────────────────────────────────────────────
# 主提取函数
# ─────────────────────────────────────────────────────────────────────────────
def extract_welds_from_pdf(pdf_path: str) -> list:
    m = re.search(r"-(BE\d+|CO\d+)_", os.path.basename(pdf_path), re.I)
    comp = m.group(1).upper() if m else os.path.splitext(os.path.basename(pdf_path))[0]

    print(f"\n{'='*60}")
    print(f"{os.path.basename(pdf_path)}  [{comp}]")

    doc  = fitz.open(pdf_path)
    rows = []

    for page_num, page in enumerate(doc, 1):
        print(f"  [页面 {page_num}/{len(doc)}] 检测截面视图...")
        views = detect_section_views(page)
        print(f"  找到 {len(views)} 个视图: {[v[0] for v in views]}")

        # 额外添加主立面视图（截面标签左侧区域，含焊缝标注）
        if views:
            leftmost_x = min(r.x0 for _, r in views)
            if leftmost_x > 200:          # 确实有左侧主视图空间
                main_rect = fitz.Rect(
                    0, page.rect.height * 0.35,
                    leftmost_x - 10, page.rect.height * 0.85,
                )
                views.insert(0, ("主视图", main_rect))
                print(f"    主视图: x=0-{leftmost_x-10:.0f}, "
                      f"y={main_rect.y0:.0f}-{main_rect.y1:.0f}")

        for section_label, rect in views:
            w_px = int(rect.width  * CROP_DPI / 72)
            h_px = int(rect.height * CROP_DPI / 72)
            print(f"  → 视图 {section_label}: 渲染 {w_px}×{h_px}px PNG，调用LLM...")

            b64      = crop_view_to_png_b64(page, rect)
            text_ctx = extract_text_in_rect(page, rect)
            print(f"    区域文字({text_ctx.count(chr(10))+1}条): {text_ctx[:200]}")
            raw = call_llm(b64, comp, section=section_label, text_ctx=text_ctx) or "[]"

            print(f"    LLM响应（{len(raw)}字）:")
            print(raw)

            view_rows = [normalize_row(r, comp) for r in parse_llm_response(raw)]
            print(f"    解析 {len(view_rows)} 条焊缝")
            rows.extend(view_rows)

    doc.close()

    # 过滤异常行并去重
    seen, deduped = set(), []
    for r in rows:
        hf_val = r.get("hf") or 0
        if not hf_val:               # hf=0 或缺失 → 跳过
            continue
        if hf_val > 25:              # hf>25 为误识别（如零件编号数字）→ 跳过
            continue
        key = (r.get("position"), r.get("hf"), r.get("length_mm"),
               r.get("part1"), r.get("part2"), r.get("annotation", ""))
        if key not in seen:
            seen.add(key)
            deduped.append(r)

    print(f"  → {len(deduped)} 条焊缝（去重后）")
    return deduped


# ─────────────────────────────────────────────────────────────────────────────
# Excel 输出
# ─────────────────────────────────────────────────────────────────────────────
def write_excel(all_results: list, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "焊缝统计_LLM"

    headers = ["序号", "位置(上/下)", "焊脚尺寸hf(mm)",
               "焊缝长度(mm)", "备注", "零件1", "零件2", "构件号"]
    ws.append(headers)

    hdr_fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(all_results, 1):
        ws.append([
            i,
            r.get("position", ""),
            r.get("hf", ""),
            r.get("length_mm", ""),
            r.get("annotation", ""),
            r.get("part1", ""),
            r.get("part2", ""),
            r.get("component", ""),
        ])

    for col, width in zip("ABCDEFGH", [6, 12, 16, 14, 14, 10, 10, 10]):
        ws.column_dimensions[col].width = width

    wb.save(output_path)
    print(f"\n已保存 → {output_path}")
    print(f"共 {len(all_results)} 条焊缝记录")


# ─────────────────────────────────────────────────────────────────────────────
# 入口
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    pdf_files = sorted(glob.glob(os.path.join(FOLDER, "*.pdf")))
    if not pdf_files:
        print(f"未找到PDF文件：{FOLDER}")
        raise SystemExit(1)

    print(f"找到 {len(pdf_files)} 个PDF文件")
    all_results = []

    for pdf_path in pdf_files:
        try:
            rows = extract_welds_from_pdf(pdf_path)
            all_results.extend(rows)
        except Exception as e:
            print(f"处理 {os.path.basename(pdf_path)} 时出错: {e}")
            traceback.print_exc()

    write_excel(all_results, OUTPUT)
